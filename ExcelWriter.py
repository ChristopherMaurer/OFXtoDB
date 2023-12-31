import openpyxl
from openpyxl import load_workbook
from copy import copy
import OFXWriter
import re
from os import path
import threading


class ConfigurationError(Exception):
    pass

# This is a replacement function called by re.sub below to parse relative cell references out of an EXCEL formula and
# replace with a cell reference to the same column and one row number greater.  This is like a Fill-Down function using
# EXCEL.  Does not support R1C1-type references nor relative named ranges (does anyone REALLY use these?)
def repl(match):
    return match.group(1) + str(1 + int(match.group(2)))


# Searching for duplicate primary keys is slow.  This attempts to speed it up by putting all primary keys into a
# dictionary structure, thus effectively making a hash index out of the primary key.  The dictionary key is the primary
# key of the EXCEL record and the item is the row number of the entry in the sheet, making it a direct access to the
# matching record.  The indexing is performed in a parallel thread, thus letting (in theory) the indexing take place at
# the same time the main program is accumulating its data records before writing them out.  Note that multithreading
# does not imply parallel processing in Python 3.11.  We'll need a future release of Python before two totally
# compute-bound threads can execute in parallel and we can reap some time savings.
def IndexOneWorksheet(tablename, ws, MapSpecs):
    pvt = threading.local()
    for pvt.colhdrs in ws.iter_rows(min_col=1, max_col=ws.max_column, min_row=1, max_row=1):
        break
    ws.FormulaCols = list(range(1,ws.max_column+1))    # Keep track of columns with Excel formulas in them
    ws.colnbrs = []
    for pvt.OFXcol in MapSpecs.Cols:
        for pvt.wscol in pvt.colhdrs:
            if pvt.OFXcol == pvt.wscol.value:
                ws.colnbrs.append(pvt.wscol.col_idx)
                ws.FormulaCols.remove(pvt.wscol.col_idx)  # Remove every column that is stored into.  What's left is columns that might have calculations
                break
        else:
            raise ConfigurationError(
                "Column {0} is not in spreadsheet for Table {1}".format(pvt.OFXcol, tablename))
    for pvt.col in ws.FormulaCols:    # Check columns left (not stored into) for the existence of a formula
        pvt.wscol = ws.cell(ws.max_row, pvt.col)
        if pvt.wscol.value is None or pvt.wscol.value[0]!="=" or ws.max_row<=1:  # ws has data + last value starts with "="
            ws.FormulaCols.remove(pvt.col)
    pvt.pkcolnbrs = []
    for pvt.PKItem in MapSpecs.PKCols:
        for pvt.wscol in pvt.colhdrs:
            if pvt.PKItem == pvt.wscol.value:
                pvt.pkcolnbrs.append(pvt.wscol.col_idx - 1)
                break
        else:
            raise ConfigurationError(
                "PK Column {0} is not in spreadsheet for Table {1}".format(pvt.PKItem, tablename))
    ws.PKIndex = {}
    pvt.rownum = 2
    for pvt.row in ws.iter_rows(min_col=1, max_col=ws.max_column, min_row=pvt.rownum, max_row=ws.max_row):
        pvt.PK = tuple(pvt.row[i].value for i in pvt.pkcolnbrs)
        ws.PKIndex[pvt.PK] = pvt.rownum
        pvt.rownum += 1
    ws.PKIndexIsReady = True


# Notice that IndexOneWorksheet can be called either as a subroutine/function or in a separate thread, making it
# possible to add some tuning logic here to avoid thread setup/takedown overhead when processing small sheets.
#  Something like: "if ws.max_row < 30: <call as a function> else: <call as a thread>".
def IndexAllWorksheets(MapSpecs, wb):
    for OFX in MapSpecs:
        for tablename in MapSpecs[OFX]:
            if tablename in wb.sheetnames:
                ws = wb[tablename]
                if (not hasattr(ws,"PKIndexIsReady")) and ws.max_row>0:
                    ws.PKIndexIsReady = False
                    spectuple = MapSpecs[OFX][tablename][0]
#                    IndexOneWorksheet(tablename, ws, spectuple)  # Old invocation - called as a function, not as a thread
                    ws.IndexThread = threading.Thread(target=IndexOneWorksheet, args=(tablename, ws, spectuple))
                    ws.IndexThread.start()  # Start up each worksheet's index build in a separate thread


class ExcelWBWriter(OFXWriter.Writer):
    def __init__(self, plist):
        fn = plist['ExcelFile'] if 'ExcelFile' in plist else None
        fn = path.expandvars(fn)
        try:
            self.wb = load_workbook(filename=fn)
        except FileNotFoundError:
            self.wb = openpyxl.Workbook()
        self.wb.workbookname = fn  # save this so when we write it out again it can be under the same name
        self.destination = re.search(r"[^\\/]+$",fn).group(0)
        super().__init__(plist)
#  Now that all mapping specs are processed and the spreadsheet is in memory, Index each sheet's primary keys
        IndexAllWorksheets(self.OFXListDict, self.wb)
        self.__anychanges = False

    def OFXListEnd(self):
        if self.curOFXList is not None:
            for EachTable in self.curOFXList:
                datatuple = self.curOFXList[EachTable][0]
                if EachTable not in self.wb.sheetnames:
                    ws = self.wb.create_sheet(title=EachTable)
                    ws.append([i for i in datatuple.Cols ])
                    for rec in self.curOFXList[EachTable][2]:
                        ws.append(rec)
                    self.__anychanges = True
                    if EachTable not in self.Stats:
                        self.Stats[EachTable] = [0, 0]  # Update statistics
                    self.Stats[EachTable][0] += len(self.curOFXList[EachTable][2])
                else:
                    ws = self.wb[EachTable]
                    if EachTable not in self.Stats:
                        self.Stats[EachTable] = [0, 0]  # Get ready to update statistics
                    origmaxrow = ws.max_row
                    if hasattr(ws,"IndexThread"): ws.IndexThread.join()   # synchronize here to the parallel thread indexing this worksheet.
                    if not ws.PKIndexIsReady:   # Thread ended, but successfully or failure?
                        raise ConfigurationError('Failed to Create Unique Index for worksheet {0}'.format(EachTable))
                    for datarow, thisPK in zip(self.curOFXList[EachTable][2], self.curOFXList[EachTable][3]):
                        self.__anychanges = True
                        destrow = ws.PKIndex[thisPK] if thisPK in ws.PKIndex else ws.max_row+1
                        for newvalue, destcol in zip(datarow,ws.colnbrs):
                            ws.cell(row=destrow, column=destcol).value = newvalue
                            if destrow > origmaxrow and destrow>1:  # Carry down formatting from above on newly-created rows
                                ws.cell(row=destrow, column=destcol).number_format\
                                    = ws.cell(row=destrow-1, column=destcol).number_format
                                ws.cell(row=destrow, column=destcol).font\
                                    = copy(ws.cell(row=destrow-1, column=destcol).font)
                        if destrow > origmaxrow:
                            for destcol in ws.FormulaCols:   # Copy-Paste the formula directly above for each new col with a formula
                                ws.cell(row=destrow, column=destcol).value = re.sub(r"(?<!\w)([a-z]{1,3})(\d+)", repl
                                            , ws.cell(row=destrow-1, column=destcol).value, flags=re.IGNORECASE)
                            self.Stats[EachTable][0] += 1
                        else:
                            self.Stats[EachTable][1] += 1

    def OFXAllDone(self):
        if self.__anychanges: self.wb.save(self.wb.workbookname)
